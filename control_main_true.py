import numpy as np
import pandas as pd
from Teams import Team, generate_lineups_six_to_four, Coach, choose_perks, TeamSeason, Captain
from contests import round_robin, alter_lineup
from colorama import Fore
from season_wipe import season_wipe
import time
from random import choice, shuffle, uniform, randint
from leagues import league_season, user_draft, player_changes, grade_players, double_elim_12, caps, captain_changes, coach_changes, the_cup
from Games import best_of, enablePrint
from stat_functions import QUERY, initiate_databases, finalize_series_data, clear_all_databases
from statistics import mean, stdev
from collections import OrderedDict, defaultdict
import re
from openpyxl import load_workbook, Workbook
import sqlite3
from switches import SEASONS, cup_frequency, activate_perks
from Players import Player


#main begins on line 387, and the first season begins on line 468


def weighted_averages(team):  # takes in a team and calculates weighted average of each stat, returns a dataframe
    #despite being called weighted_averages, this produces a full dataframe for a team season

    final_dict = {'Team': team.name, 'Power': None, 'Attack Damage': None, 'Attack Speed': None, 'Critical %': None,
                  'Critical X': None, 'Health': None, 'Spawn Time': 0, 'Lineup Wins': team.wins,
                  'Lineup Losses': team.losses
        , 'Match Wins': team.match_wins, 'Match Losses': team.match_losses, 'Match Draws': team.match_draws}

    power_list = [player.power for player in team.players]
    total_power = (power_list[0] * 9) + (power_list[1] * 7) + (power_list[2] * 6) + (power_list[3] * 6) + (
                power_list[4] * 4) + (power_list[5] * 4)
    final_dict['Power'] = round((total_power / 36), 2)

    atk_dmg_list = [player.atk_dmg for player in team.players]
    total_atk_dmg = (atk_dmg_list[0] * 9) + (atk_dmg_list[1] * 7) + (atk_dmg_list[2] * 6) + (atk_dmg_list[3] * 6) + (
                atk_dmg_list[4] * 4) + (atk_dmg_list[5] * 4)
    final_dict['Attack Damage'] = round((total_atk_dmg / 36), 2)

    atk_spd_list = [player.atk_spd for player in team.players]
    total_atk_spd = (atk_spd_list[0] * 9) + (atk_spd_list[1] * 7) + (atk_spd_list[2] * 6) + (atk_spd_list[3] * 6) + (
                atk_spd_list[4] * 4) + (atk_spd_list[5] * 4)
    final_dict['Attack Speed'] = round((total_atk_spd / 36), 2)

    crit_pct_list = [player.crit_pct for player in team.players]
    total_crit_pct = (crit_pct_list[0] * 9) + (crit_pct_list[1] * 7) + (crit_pct_list[2] * 6) + (
                crit_pct_list[3] * 6) + (crit_pct_list[4] * 4) + (crit_pct_list[5] * 4)
    final_dict['Critical %'] = round((total_crit_pct / 36), 5)

    crit_x_list = [player.crit_x for player in team.players]
    total_crit_x = (crit_x_list[0] * 9) + (crit_x_list[1] * 7) + (crit_x_list[2] * 6) + (crit_x_list[3] * 6) + (
                crit_x_list[4] * 4) + (crit_x_list[5] * 4)
    final_dict['Critical X'] = round((total_crit_x / 36), 3)

    health_list = [player.max_health for player in team.players]
    total_health = (health_list[0] * 9) + (health_list[1] * 7) + (health_list[2] * 6) + (health_list[3] * 6) + (
                health_list[4] * 4) + (health_list[5] * 4)
    final_dict['Health'] = round((total_health / 36), 3)

    spawn_list = [player.spawn_time for player in team.players]
    total_spawn = (spawn_list[0] * 9) + (spawn_list[1] * 7) + (spawn_list[2] * 6) + (spawn_list[3] * 6) + (
                spawn_list[4] * 4) + (spawn_list[5] * 4)
    final_dict['Spawn Time'] = round((total_spawn / 36), 2)

    return pd.DataFrame([final_dict])

def sort_players_by_xWAR(team):
    team.players.sort(key=lambda pl: pl.xWAR, reverse=True)

def sort_all_players(leagues, manual):
    #teams have a boolean called mine
    def move_player(old, new, swap_team):
        swap_team.players[old], swap_team.players[new] = swap_team.players[new], swap_team.players[old]

    for league in leagues:
        for team in league:
            grade_players(team.players, is_team=True)
            if team.mine:
                enablePrint()
                if manual:
                    print(f"Here is the current order of your players ({team.name}):" + Fore.RESET)
                    i = 0
                    for player in team.players:
                        print(f"Slot {i}, {player}")
                        i += 1
                    while True:
                        user_choice = input(Fore.BLUE + "Would you like to move anything? If not, press N.\n"
                                                        "To move a player, press their Slot Number.")
                        if user_choice == 'N' or user_choice == 'n':
                            break
                        else:
                            user_choice = int(user_choice)
                        old_slot = user_choice
                        new_slot = int(input("What will their new slot be?"))
                        move_player(old_slot, new_slot, team)
                        print(Fore.RESET + "Here is the new order of your players:")
                        i = 0
                        for player in team.players:
                            print(f"Slot {i}, {player}")
                            i += 1
            team.lineups = generate_lineups_six_to_four(team.players, team.team_coach)



def write_to_file(filename=None, words=None, mode='w', error=False):
    if error:
        filename = 'error_output'
        mode = 'a'

    with open(filename, mode) as f:
        f.write(words + '\n')


def clean_file(file_path):
    with open(file_path, 'r+') as file:
        lines = file.readlines()
        file.seek(0)
        file.truncate()

        delete_next_line = False

        for line in lines:
            # Check if the line contains a colon
            if ':' in line:
                delete_next_line = True
                continue

            # Check if the line should be deleted based on the previous line
            if delete_next_line:
                delete_next_line = False
                continue

            # Check if the line is empty
            if line.strip() == '':
                continue

            # Write the line back to the file
            file.write(line)

def format_champion_line(line):
    # Define a regex pattern to match the line format
    if "Regional" in line:

        pattern = re.compile(r'^S(\d+)\s(.+?)\sRegional\sChampion:\s(.+)$')

        # Use the regex pattern to match and extract groups
        match = pattern.match(line)

        # If the line matches the pattern, format it as required
        if match:
            seed = match.group(1)
            region = match.group(2)
            champion = match.group(3)
            formatted_line = f"{champion} (S{seed} {region} Regional Champion)"
            return formatted_line
        else:
            # If the line doesn't match the expected format, return the original line
            return line.strip()
    else:
        # Define a regex pattern to match the required line format
        pattern = re.compile(r'^S(\d+)\s(.+?)\sChampion:\s(.+)$')

        # Use the regex pattern to match and extract groups
        match = pattern.match(line)

        # If the input string matches the format, format it as required
        if match:
            seed = match.group(1)
            champion = match.group(3)
            formatted_string = f"{champion} (S{seed} {champion} Champion)"
            return formatted_string
        else:
            # If the input string doesn't match the expected format, return the original string
            return line.strip()

def remove_duplicates_ordered(upl_standings):
    seen_teams = OrderedDict()
    unique_standings = []

    for team in upl_standings:
        if team not in seen_teams:
            unique_standings.append(team)
            seen_teams[team] = None

    return unique_standings

def extract_high_seed(text):
    int_values = [int(match.group()) for match in re.finditer(r'\b\d+\b', text)][:2]
    high_seed = min(int_values)
    return high_seed


def clear_file(filename, excel=False,sheets='all'):
    if not excel:
        if isinstance(filename,list):
            for file in filename:
                with open(file, 'w') as c:
                    c.write('')
        else:
            with open(filename, 'w') as c:
                c.write('')
    else:
        workbook = load_workbook(filename)

        # Loop through every sheet in the workbook
        for sheet in (workbook.sheetnames if sheets=='all' else [sheets] if isinstance(sheets, str) else sheets):
            worksheet = workbook[sheet]

            worksheet.delete_rows(1, worksheet.max_row)

        # Save the cleared workbook
        workbook.save(filename)

def create_teams(count,region='None',season_count=0, draft=True, allow_duplicate_names=False):
    TEAMS = []
    for i in range(count):
        temp = Team(region,season_count=season_count,allow_duplicate_names=allow_duplicate_names)
        TEAMS.append(temp)
    if draft:
        user_draft(TEAMS, season_count, is_regional=True, draft_name=f"{region} Preliminary Draft - Round 1")
        user_draft(TEAMS, season_count, is_regional=True, draft_name=f"{region} Preliminary Draft - Round 2")
        user_draft(TEAMS, season_count, is_regional=True, draft_name=f"{region} Preliminary Draft - Round 3")
    return TEAMS

def ordinal_string(n: int) -> str:
    if 10 <= n % 100 < 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

def main():
    from switches import jackson_playing, use_saved

    import warnings
    warnings.simplefilter(action='ignore', category=FutureWarning)

    start = time.time()
    # file clearing:
    clear_file(['my_teams', 'error_output', 'upsets',
                 'team_coach_data', 'execution_time', 'off_season_report', 'my_team_playerstats',
                'my_team_results', 'comebacks', 'captain_off_season_report'])



    all_teams = []
    #this is the only list which should contain every single season list, from every single season, period


    my_teams = []
    jackson_teams = []

    if use_saved:
        print(Fore.RED + "This feature does not work. Starting new iteration.")
        use_saved = False
    if not use_saved:
        season_count = 0
        uni_teams = create_teams(26, "Universal", season_count=season_count)

        for team in uni_teams:
            team.history[season_count] = ""
        upl_standings = league_season(uni_teams, use_saved=False, season_count=0, region="Universal")
        dw_sc = choice([True, False])
        ds_wof = choice([True, False])
        iw_cl = choice([True, False])
        hc_sh = choice([True, False])


        dw_teams = create_teams((21 if dw_sc or jackson_playing else 22), "Darkwing", season_count=season_count)
        sc_teams = create_teams((21 if not dw_sc or jackson_playing else 22), "Shining-Core", season_count=season_count)
        ds_teams = create_teams((21 if ds_wof or jackson_playing else 22), "Diamond-Sea", season_count=season_count)
        wof_teams = create_teams((21 if not ds_wof or jackson_playing else 22), "Web-of-Nations", season_count=season_count)
        iw_teams = create_teams((21 if iw_cl or jackson_playing else 22), "Ice-Wall", season_count=season_count)
        cl_teams = create_teams((21 if not iw_cl or jackson_playing else 22), "Candyland", season_count=season_count)
        hc_teams = create_teams((21 if hc_sh or jackson_playing else 22), "Hell's-Circle", season_count=season_count)
        sh_teams = create_teams((21 if not hc_sh or jackson_playing else 22), "Steel-Heart", season_count=season_count)


        my_team1 = Team(("Darkwing" if dw_sc else "Shining-Core"), mine=True, season_count=season_count)
        jackson_team1 = Team(("Darkwing" if not dw_sc else "Shining-Core"), mine=True, season_count=season_count,jackson=True)

        my_team2 = Team(("Diamond-Sea" if ds_wof else "Web-of-Nations"), mine=True, season_count=season_count)
        jackson_team2 = Team(("Diamond-Sea" if not ds_wof else "Web-of-Nations"), mine=True, season_count=season_count,jackson=True)

        my_team3 = Team(("Ice-Wall" if iw_cl else "Candyland"), mine=True, season_count=season_count)
        jackson_team3 = Team(("Ice-Wall" if not iw_cl else "Candyland"), mine=True, season_count=season_count,
                             jackson=True)

        my_team4 = Team(("Hell's-Circle" if hc_sh else "Steel-Heart"), mine=True, season_count=season_count)
        jackson_team4 = Team(("Hell's-Circle" if not hc_sh else "Steel-Heart"), mine=True, season_count=season_count,
                             jackson=True)

        if not jackson_playing:
            if dw_sc:
                dw_teams.append(my_team1)
            else:
                sc_teams.append(my_team1)

            if ds_wof:
                ds_teams.append(my_team2)
            else:
                wof_teams.append(my_team2)

            if iw_cl:
                iw_teams.append(my_team3)
            else:
                cl_teams.append(my_team3)

            if hc_sh:
                hc_teams.append(my_team4)
            else:
                sh_teams.append(my_team4)
            jackson_teams = []
        else:
            if dw_sc:
                dw_teams.append(my_team1)
                sc_teams.append(jackson_team1)
            else:
                sc_teams.append(my_team1)
                dw_teams.append(jackson_team1)

            if ds_wof:
                ds_teams.append(my_team2)
                wof_teams.append(jackson_team2)
            else:
                wof_teams.append(my_team2)
                ds_teams.append(jackson_team2)

            if iw_cl:
                iw_teams.append(my_team3)
                cl_teams.append(jackson_team3)
            else:
                cl_teams.append(my_team3)
                iw_teams.append(jackson_team3)

            if hc_sh:
                hc_teams.append(my_team4)
                sh_teams.append(jackson_team4)
            else:
                sh_teams.append(my_team4)
                hc_teams.append(jackson_team4)
            jackson_teams = [jackson_team1,jackson_team2,jackson_team3,jackson_team4]

        my_teams.extend([my_team1,my_team2,my_team3,my_team4])


        for i in range((len(my_teams))):
            my_team = my_teams[i]
            my_team.make_mine(my_team.name)
            ch_slots_amped = choice([[0,5], [1,4],[0,4,5], [1,4,5], [1,3,5]])
            ch_att_amped = choice([["Critical Chance", round(uniform(0.021,0.033), 3)], ["Attack Damage", randint(3,6)]])
            ch_slot_effect = [ch_slots_amped]
            ch_slot_effect.extend(ch_att_amped)
            choward_coach = Coach(region=my_team.region, fixed_slot_effect=ch_slot_effect,fixed_name="Carter Howard")
            my_team.change_coach(choward_coach)
            print('\n' + Fore.BLUE + f"{my_team.name} is your team. Coach: {str(choward_coach)}\n {str(my_team.captain)}")
            my_team.print_roster(first=True)
        if jackson_playing:
            for i in range(len(jackson_teams)):
                j_team = jackson_teams[i]
                jk_slots_amped = choice([[0], [1]])
                jk_att_amped = ["Power", round(uniform(0.839, 0.99), 2)]
                jk_slot_effect = [jk_slots_amped]
                jk_slot_effect.extend(jk_att_amped)
                jkeefe_coach = Coach(region=j_team.region, fixed_slot_effect=jk_slot_effect,fixed_name="Jackson Keefe")
                j_team.change_coach(jkeefe_coach)
                print('\n' + Fore.BLUE + f"{j_team.name} is your team. Coach: {str(jkeefe_coach)}\n {str(j_team.captain)}")
                j_team.print_roster(first=True)



        print(Fore.RESET)

    end_season_times = [float(0) for _ in range(season_count + SEASONS + 1)]

    def regional_leagues(dw_teams,sc_teams,ds_teams,wof_teams,iw_teams,cl_teams,hc_teams,sh_teams,season_count):
            pre_qualif_tournament = []
            last_stand_tournament = {'5 Seeds' : [], '6 Seeds' : [], '7 Seeds' : [], '8 Seeds' : []}

            print(Fore.GREEN + "DARKWING REGION, " + Fore.RESET, end='')
            dw_teams = league_season(dw_teams, False, season_count=season_count, final_reversed=False,
                                     region='Darkwing Regional')
            dw_qualified = dw_teams[:8]


            print(Fore.GREEN + "SHINING CORE REGION, " + Fore.RESET, end='')
            sc_teams = league_season(sc_teams, False, season_count=season_count, final_reversed=False,
                                     region='Shining-Core Regional')
            sc_qualified = sc_teams[:8]

            print(Fore.GREEN + "DIAMOND SEA REGION, " + Fore.RESET, end='')
            ds_teams = league_season(ds_teams, False, season_count=season_count, final_reversed=False,
                                     region='Diamond-Sea Regional')
            ds_qualified = ds_teams[:8]

            print(Fore.GREEN + "WEB OF NATIONS, " + Fore.RESET, end='')
            wof_teams = league_season(wof_teams, False, season_count=season_count, final_reversed=False,
                                      region='Web-of-Nations Regional')
            wof_qualified = wof_teams[:8]

            print(Fore.GREEN + "ICE WALL REGION, " + Fore.RESET, end='')
            iw_teams = league_season(iw_teams, False, season_count=season_count, final_reversed=False,
                                     region='Ice-Wall Regional')
            iw_qualified = iw_teams[:8]

            print(Fore.GREEN + "CANDYLAND REGION, " + Fore.RESET, end='')
            cl_teams = league_season(cl_teams, False, season_count=season_count, final_reversed=False,
                                     region='Candyland Regional')
            cl_qualified = cl_teams[:8]

            print(Fore.GREEN + "HELL'S CIRCLE, " + Fore.RESET, end='')
            hc_teams = league_season(hc_teams, False, season_count=season_count, final_reversed=False,
                                     region="Hell's-Circle Regional")
            hc_qualified = hc_teams[:8]

            print(Fore.GREEN + "STEEL HEART REGION, " + Fore.RESET, end='')
            sh_teams = league_season(sh_teams, False, season_count=season_count, final_reversed=False,
                                     region="Steel-Heart Regional")
            sh_qualified = sh_teams[:8]

            dw_champ = dw_qualified[0]
            regional_champs.append(dw_champ)
            sc_champ = sc_qualified[0]
            regional_champs.append(sc_champ)
            ds_champ = ds_qualified[0]
            regional_champs.append(ds_champ)
            wof_champ = wof_qualified[0]
            regional_champs.append(wof_champ)
            iw_champ = iw_qualified[0]
            regional_champs.append(iw_champ)
            cl_champ = cl_qualified[0]
            regional_champs.append(cl_champ)
            hc_champ = hc_qualified[0]
            regional_champs.append(hc_champ)
            sh_champ = sh_qualified[0]
            regional_champs.append(sh_champ)

            # trans_region = ['Darkwing', 'Shining-Core', 'Diamond-Sea', 'Web-of-Nations', 'Ice-Wall', 'Candyland', "Hell's-Circle", 'Steel-Heart']
            count = 0
            for region in [dw_qualified, sc_qualified, ds_qualified, wof_qualified, iw_qualified, cl_qualified,
                           hc_qualified, sh_qualified]:
                for number in [1, 2, 3]:
                    pre_qualif_tournament.append(region[number])

                last_stand_tournament['5 Seeds'].append(region[4])
                last_stand_tournament['6 Seeds'].append(region[5])
                last_stand_tournament['7 Seeds'].append(region[6])
                last_stand_tournament['8 Seeds'].append(region[7])
                count += 1

            for region in [dw_qualified, sc_qualified, ds_qualified, wof_qualified, iw_qualified, cl_qualified,
                            hc_qualified, sh_qualified]:
                season_wipe(region)

            if len(last_stand_tournament['8 Seeds']) < 8:
                enablePrint()
                print(Fore.RED + "NOT ENOUGH 8 SEEDS!" + Fore.RESET)
                write_to_file(error=True, words="NOT ENOUGH 8 SEEDS!")
                time.sleep(1)
            if len(last_stand_tournament['7 Seeds']) < 8:
                enablePrint()
                print(Fore.RED + "NOT ENOUGH 7 SEEDS!" + Fore.RESET)
                write_to_file(error=True, words="NOT ENOUGH 7 SEEDS!")
                time.sleep(1)
            if len(last_stand_tournament['6 Seeds']) < 8:
                enablePrint()
                print(Fore.RED + "NOT ENOUGH 6 SEEDS!" + Fore.RESET)
                write_to_file(error=True, words="NOT ENOUGH 6 SEEDS!")
                time.sleep(1)
            if len(last_stand_tournament['5 Seeds']) < 8:
                enablePrint()
                print(Fore.RED + "NOT ENOUGH 5 SEEDS!" + Fore.RESET)
                write_to_file(error=True, words="NOT ENOUGH 6 SEEDS!")
                time.sleep(1)
            return regional_champs, pre_qualif_tournament, last_stand_tournament
    another = 'y'
    relegated = {} # int keys, team object values
    promoted = {}

    for num in range(SEASONS):

        if num >= 1:
            write_to_file(filename='my_team_results', words="\n",
                          mode='a', error=False)

        if activate_perks:
            for team in my_teams:
                choose_perks(team)
            if jackson_playing:
                for team in jackson_teams:
                    choose_perks(team)

        season_count += 1

        write_to_file('comebacks', words=f"SEASON {season_count}", mode='a')

        total_stats = {'Power': 0, 'Attack Damage': 0, 'Attack Speed': 0, 'Critical %': 0, 'Critical X': 0,
                       'Spawn Time': 0, 'Health': 0, 'Age': 0}
        all_players = []
        with open("rosters", 'a') as w:
            w.write(f"SEASON NO. {season_count}\n")
            trans_region = ['Universal', 'Darkwing', 'Shining-Core', 'Diamond-Sea', 'Web-of-Nations', 'Ice-Wall',
                            'Candyland', "Hell's-Circle", 'Steel-Heart']
            trans_i = -1
            for league in [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
                trans_i += 1
                for team in league:
                    team.region = trans_region[trans_i]
                    for player in team.players:
                        total_stats['Power'] += player.power
                        total_stats['Attack Damage'] += player.atk_dmg
                        total_stats['Attack Speed'] += player.atk_spd
                        total_stats['Critical %'] += player.crit_pct
                        total_stats['Critical X'] += player.crit_x
                        total_stats['Spawn Time'] += player.spawn_time
                        total_stats['Health'] += player.max_health
                        total_stats['Age'] += player.age

                        all_players.append(player)
            grade_players(all_players, is_team=True)


        relegated[season_count] = []
        promoted[season_count] = []

        sort_all_players([uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams], manual=False)

        with open("my_rosters", 'w') as file:
            file.write(f"!!!!!!_SEASON {season_count}_!!!!!!\n")
        for league in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
            season_wipe(league)
            for team in league:
                team.fired_coach_this_season = False
                team.history[season_count] = ""

                if team.mine or team.jackson:
                    team.print_roster(first=True)
        if season_count != 0:
            season_wipe(uni_teams)
            for team in uni_teams:
                team.history[season_count] = ""

        uni_qualif_g1 = []
        uni_qualif_g2 = []
        regional_champs = []
        last_stand = []
        regional_champs, pqt, last_stand = regional_leagues(dw_teams,sc_teams,ds_teams,wof_teams,iw_teams,cl_teams,hc_teams,sh_teams,season_count=season_count)

        cup_teams = []
        if season_count % cup_frequency == 0:
            for league in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams, upl_standings]:
                for team in league:
                    cup_teams.append(team)
            the_cup(cup_teams)


        last_stand_groups = {key: [] for key in range(8)}
        shuffle(last_stand['5 Seeds'])
        shuffle(last_stand['6 Seeds'])
        shuffle(last_stand['7 Seeds'])
        shuffle(last_stand['8 Seeds'])


        for i in range(8):
            temp7_processed = False
            temp8_processed = False
            temp5_processed = False

            other_region = ""
            temp_6 = last_stand['6 Seeds'][i]
            temp_6.accolades['Last-Stand'] += 1
            last_stand_groups[i].append(temp_6)

            for temp_7 in last_stand['7 Seeds']:
                if temp_7.played_region[season_count] != temp_6.played_region[season_count]:
                    other_region = temp_7.played_region[season_count]
                    temp_7.accolades['Last-Stand'] += 1
                    last_stand_groups[i].append(temp_7)
                    last_stand['7 Seeds'].remove(temp_7)
                    temp7_processed = True
                    break

            if not temp7_processed:
                write_to_file(error=True, words=f"temp7 failsafe triggered. Length {len(last_stand['7 Seeds'])}")
                temp_7 = choice(last_stand['7 Seeds'])
                other_region = temp_7.played_region[season_count]
                temp_7.accolades['Last-Stand'] += 1
                last_stand_groups[i].append(temp_7)
                last_stand['7 Seeds'].remove(temp_7)

            for temp_8 in last_stand['8 Seeds']:
                if temp_8.played_region[season_count] not in [temp_6.played_region[season_count], other_region]:
                    temp_8.accolades['Last-Stand'] += 1
                    last_stand_groups[i].append(temp_8)
                    last_stand['8 Seeds'].remove(temp_8)
                    temp8_processed = True
                    break

            if not temp8_processed:
                write_to_file(error=True, words=f"temp8 failsafe triggered. Length {len(last_stand['8 Seeds'])}")
                temp_8 = choice(last_stand['8 Seeds'])
                temp_8.accolades['Last-Stand'] += 1
                last_stand_groups[i].append(temp_8)
                last_stand['8 Seeds'].remove(temp_8)

            for temp_5 in last_stand['5 Seeds']:
                if temp_5.played_region[season_count] not in [temp_6.played_region[season_count], other_region]:
                    temp_5.accolades['Last-Stand'] += 1
                    last_stand_groups[i].append(temp_5)
                    last_stand['5 Seeds'].remove(temp_5)
                    temp5_processed = True
                    break

            if not temp5_processed:
                write_to_file(error=True, words=f"temp5 failsafe triggered. Length {len(last_stand['8 Seeds'])}")
                temp_5 = choice(last_stand['5 Seeds'])
                temp_5.accolades['Last-Stand'] += 1
                last_stand_groups[i].append(temp_5)
                last_stand['5 Seeds'].remove(temp_5)

        last_stand_start_time = time.time()
        index = 0
        ls_group_names = ["A", 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for key in last_stand_groups.keys():
            print(Fore.GREEN + f"LAST STAND, GROUP {ls_group_names[index]}: " + Fore.RESET, end='')
            index+=1
            temp_standings = round_robin(last_stand_groups[key], 24, 4, print_region_seed=True,cyan_seeds=[0,1],
                                                       red_seeds=[2,3], print_by_round=False)

            temp_standings[0].history[season_count] += f" 1st in Last Stand Group {key} -> Pre-Qualifying Group."
            pqt.append(temp_standings[0])
            temp_standings[0].team_seasons[season_count].last_stand = 2

            temp_standings[1].history[season_count] += f" 2nd in Last Stand Group {key} -> Pre-Qualifying Group."
            pqt.append(temp_standings[1])
            temp_standings[1].team_seasons[season_count].last_stand = 2

            temp_standings[2].history[season_count] += f" Eliminated from Last Stand. (3rd in Group {key})"
            temp_standings[2].team_seasons[season_count].last_stand = 1

            temp_standings[3].history[season_count] += f" Eliminated from Last Stand. (4th in Group {key})"
            temp_standings[3].team_seasons[season_count].last_stand = 1

            season_wipe(temp_standings)
        last_stand_end_time = time.time()
        with open('execution_time', 'a') as file:
            file.write(f"Season {season_count} Last Stand Execution Time: {round((last_stand_end_time - last_stand_start_time) / 60)} minutes, {round(((last_stand_end_time - last_stand_start_time) % 60), 2)} seconds.\n")

        for team in pqt:
            team.accolades['Pre-Qualifying'] += 1

        def parse_region_seed(a_team):
            # Split the region_seed into letters and integer
            region, seed = a_team.region_seed.split('-')
            return region, int(seed)

        def create_pre_qualifying_groups(pq_teams):
            group_a, group_b, group_c, group_d, group_e, group_f, group_g, group_h = [], [], [], [], [], [], [], []

            team_by_seed = {
                2: [], 3: [], 4: [], "Last-Stand" : [],
            }

            for pq_team in pq_teams:
                region, seed = parse_region_seed(pq_team)
                if seed == 2:
                    team_by_seed[2].append(pq_team)
                    pq_team.team_seasons[season_count].last_stand = 3
                elif seed == 3:
                    team_by_seed[3].append(pq_team)
                    pq_team.team_seasons[season_count].last_stand = 3
                elif seed == 4:
                    team_by_seed[4].append(pq_team)
                    pq_team.team_seasons[season_count].last_stand = 3
                elif seed in [5,6,7,8]:
                    team_by_seed["Last-Stand"].append(pq_team)

            if len(team_by_seed[2]) != 8:
                print(Fore.RED + f"2 Seed PQ Group Error: Length {len(team_by_seed[2])}" + Fore.RESET)
            if len(team_by_seed[3]) != 8:
                print(Fore.RED + f"3 Seed PQ Group Error: Length {len(team_by_seed[3])}" + Fore.RESET)
            if len(team_by_seed[4]) != 8:
                print(Fore.RED + f"4 Seed PQ Group Error: Length {len(team_by_seed[4])}" + Fore.RESET)
            if len(team_by_seed["Last-Stand"]) != 16:
                print(Fore.RED + f"Last Stand PQ Group Error: Length {len(team_by_seed[4])}" + Fore.RESET)



            groups = [group_a, group_b, group_c, group_d, group_e, group_f, group_g, group_h]
            two_seed, three_seed, four_seed, ls_team1, ls_team2 = None, None, None, None, None
            #two_region, three_region, four_region, ls1_region, ls2_region = None, None, None, None, None
            three_found, four_found, ls1_found, ls2_found = False, False, False, False
            for g_i in range(8):
                three_found, four_found, ls1_found, ls2_found = False, False, False, False

                two_seed = choice(team_by_seed[2])
                team_by_seed[2].remove(two_seed)
                two_region = parse_region_seed(two_seed)[0]

                for t in team_by_seed[3]:
                    if parse_region_seed(t)[0] != two_region:
                        three_seed = t
                        team_by_seed[3].remove(three_seed)
                        three_found = True
                        break
                if not three_found:
                    three_seed = choice(team_by_seed[3])
                    team_by_seed[3].remove(three_seed)
                three_region = parse_region_seed(three_seed)[0]

                for t in team_by_seed[4]:
                    if parse_region_seed(t)[0] not in [two_region,three_region]:
                        four_seed = t
                        team_by_seed[4].remove(four_seed)
                        four_found = True
                        break
                if not four_found:
                    four_seed = choice(team_by_seed[4])
                    team_by_seed[4].remove(four_seed)
                four_region = parse_region_seed(four_seed)[0]

                for t in team_by_seed["Last-Stand"]:
                    if parse_region_seed(t)[0] not in [two_region,three_region,four_region]:
                        ls_team1 = t
                        team_by_seed["Last-Stand"].remove(ls_team1)
                        ls1_found = True
                        break
                if not ls1_found:
                    ls_team1 = choice(team_by_seed["Last-Stand"])
                    team_by_seed["Last-Stand"].remove(ls_team1)

                ls1_region = parse_region_seed(ls_team1)[0]

                for t in team_by_seed["Last-Stand"]:
                    if parse_region_seed(t)[0] not in [two_region,three_region,four_region,ls1_region]:
                        ls_team2 = t
                        team_by_seed["Last-Stand"].remove(ls_team2)
                        ls2_found = True
                        break
                if not ls2_found:
                    ls_team2 = choice(team_by_seed["Last-Stand"])
                    team_by_seed["Last-Stand"].remove(ls_team2)


                groups[g_i].extend([two_seed, three_seed, four_seed, ls_team1, ls_team2])
            #I'm changing this for larger groups so that placement in the universal league is not as dependent on group luck
            group_1 = group_a + group_b
            group_2 = group_c + group_d
            group_3 = group_e + group_f
            group_4 = group_g + group_h

            return group_1, group_2, group_3, group_4

        group_one, group_two, group_three, group_four = create_pre_qualifying_groups(pqt)

        print(Fore.GREEN + "PRE-QUALIFYING, GROUP ONE: " + Fore.RESET, end='')
        group_one_standings = round_robin(group_one, r=15, qualify_range=len(group_one), print_region_seed=True,
                                        cyan_seeds=[0,1,2,3,4,5], red_seeds=[6,7,8,9], print_by_round=False)
        group_one_qualif = group_one_standings[:6]
        print(Fore.GREEN + "PRE-QUALIFYING, GROUP TWO: " + Fore.RESET, end='')
        group_two_standings = round_robin(group_two, r=15, qualify_range=len(group_two), print_region_seed=True,
                                        cyan_seeds=[0,1,2,3,4,5], red_seeds=[6,7,8,9], print_by_round=False)
        group_two_qualif = group_two_standings[:6]
        print(Fore.GREEN + "PRE-QUALIFYING, GROUP THREE: " + Fore.RESET, end='')
        group_three_standings = round_robin(group_three, r=15, qualify_range=len(group_three), print_region_seed=True,
                                        cyan_seeds=[0,1,2,3,4,5], red_seeds=[6,7,8,9], print_by_round=False)
        group_three_qualif = group_three_standings[:6]
        print(Fore.GREEN + "PRE-QUALIFYING, GROUP FOUR: " + Fore.RESET, end='')
        group_four_standings = round_robin(group_four, r=15, qualify_range=len(group_four), print_region_seed=True,
                                        cyan_seeds=[0,1,2,3,4,5], red_seeds=[6,7,8,9], print_by_round=False)
        group_four_qualif = group_four_standings[:6]

        i=0
        for competitor in group_one:
            i += 1
            if competitor in group_one_qualif:
                uni_qualif_g1.append(competitor)
                competitor.team_seasons[season_count].pre_qualifying = 2
                if i == 1:
                    competitor.history[season_count] += f"Won PQ Group One -> UNI Qualifying."
                else:
                    competitor.history[season_count] += f"{ordinal_string(i)} in PQ Group One -> UNI Qualifying."
            else:
                competitor.history[season_count] += f"Eliminated in PQ ({ordinal_string(i)} in Group One)"
                competitor.team_seasons[season_count].pre_qualifying = 1
        i = 0
        for competitor in group_two:
            i += 1
            if competitor in group_two_qualif:
                uni_qualif_g1.append(competitor)
                competitor.team_seasons[season_count].pre_qualifying = 2
                if i == 1:
                    competitor.history[season_count] += f"Won PQ Group Two -> UNI Qualifying."
                else:
                    competitor.history[season_count] += f"{ordinal_string(i)} in PQ Group Two -> UNI Qualifying."
            else:
                competitor.history[season_count] += f"Eliminated in PQ ({ordinal_string(i)} in Group Two)"
                competitor.team_seasons[season_count].pre_qualifying = 1
        i = 0
        for competitor in group_three:
            i += 1
            if competitor in group_three_qualif:
                uni_qualif_g2.append(competitor)
                competitor.team_seasons[season_count].pre_qualifying = 2
                if i == 1:
                    competitor.history[season_count] += f"Won PQ Group Three -> UNI Qualifying."
                else:
                    competitor.history[season_count] += f"{ordinal_string(i)} in PQ Group Three -> UNI Qualifying."
            else:
                competitor.history[season_count] += f"Eliminated in PQ ({ordinal_string(i)} in Group Three)"
                competitor.team_seasons[season_count].pre_qualifying = 1
        i = 0
        for competitor in group_four:
            i+=1
            if competitor in group_four_qualif:
                uni_qualif_g2.append(competitor)
                competitor.team_seasons[season_count].pre_qualifying = 2
                if i == 1:
                    competitor.history[season_count] += f"Won PQ Group Four -> UNI Qualifying."
                else:
                    competitor.history[season_count] += f"{ordinal_string(i)} in PQ Group Four -> UNI Qualifying."
            else:
                competitor.history[season_count] += f"Eliminated in PQ ({ordinal_string(i)} in Group Four)"
                competitor.team_seasons[season_count].pre_qualifying = 1



        if len(upl_standings) < 30:
            drop_range1 = [0, 2, 5]
            drop_range2 = [1, 3, 4]
        else:
            drop_range1 = [0, 2, 5, 6, 9, 11, 12,14]
            drop_range2 = [1, 3, 4, 7, 8, 10, 13,15]

        for i in drop_range1:
            uni_qualif_g1.append(upl_standings[i])
            uni_teams.remove(upl_standings[i])
            relegated[season_count].append(upl_standings[i])
            upl_standings[i].team_seasons[season_count].region_started = "Universal Qualifying"

        for i in drop_range2:
            uni_qualif_g2.append(upl_standings[i])
            uni_teams.remove(upl_standings[i])
            relegated[season_count].append(upl_standings[i])
            upl_standings[i].team_seasons[season_count].region_started = "Universal Qualifying"

        for team in uni_qualif_g1:
            team.accolades['Universal-Qualifying'] += 1
        for team in uni_qualif_g2:
            team.accolades['Universal-Qualifying'] += 1

        uni_qualif_start_time = time.time()

        for i in range(8):
            if i % 2 == 0:
                uni_qualif_g1.append(regional_champs[i])
            elif i % 2 == 1:
                uni_qualif_g2.append(regional_champs[i])
        for group in [uni_qualif_g1, uni_qualif_g2]:
            season_wipe(group)

        print(Fore.GREEN + "GROUP 1 QUALIFYING ROUND, " + Fore.RESET, end='')
        g1_pre_advance = round_robin(uni_qualif_g1, 6, len(uni_qualif_g1), cyan_seeds=[0,1,2,3,4,5], yellow_seeds=[6,7,8,9], red_seeds=[r for r in range(10,len(uni_qualif_g1))])
        print(Fore.GREEN + "GROUP 2 QUALIFYING ROUND, " + Fore.RESET, end='')
        g2_pre_advance = round_robin(uni_qualif_g2, 6, len(uni_qualif_g2), cyan_seeds=[0,1,2,3,4,5], yellow_seeds=[6,7,8,9], red_seeds=[r for r in range(10,len(uni_qualif_g2))])
        uni_teams.reverse()
        g1_advance = [None for _ in range(24)]
        g2_advance = [None for _ in range(24)]

        void_draft = []

        for i in range(6): #seeds 1 through 6 who clinch universal league
            g1_advance[i] = g1_pre_advance[i]
            uni_teams.append(g1_advance[i])
            g1_advance[i].team_seasons[season_count].uni_qualifying = 2
            g1_pre_advance[i].history[season_count] += f" {ordinal_string(i+1)} in Group 1 Qualifying. -> UNI League."

            g2_advance[i] = g2_pre_advance[i]
            uni_teams.append(g2_advance[i])
            g2_advance[i].team_seasons[season_count].uni_qualifying = 2
            g2_pre_advance[i].history[season_count] += f" {ordinal_string(i+1)} in Group 2 Qualifying. -> UNI League."
        for i in [6,7,8,9]: #7, 8, 9, 10, seeds going to play-in
            g1_advance[i] = g1_pre_advance[i]
            g1_pre_advance[i].qualifying_group = 1
            g1_pre_advance[i].history[season_count] += f" {ordinal_string(i + 1)} in Group 1 Qualifying. -> Universal Play-In. "

            g2_advance[i] = g2_pre_advance[i]
            g2_pre_advance[i].qualifying_group = 2
            g2_pre_advance[i].history[season_count] += f" {ordinal_string(i + 1)} in Group 2 Qualifying. -> Universal Play-In. "
        for i in range(10,len(g1_pre_advance)): #11 seed and below eliminated
            void_draft.append(g1_pre_advance[i])
            void_draft.append(g2_pre_advance[i])

            g1_pre_advance[i].history[season_count] += f" {ordinal_string(i+1)} in Group 1 Qualifying. Failed to qualify."
            g1_pre_advance[i].team_seasons[season_count].uni_qualifying = 1

            g2_pre_advance[i].history[season_count] += f" {ordinal_string(i+1)} in Group 2 Qualifying. Failed to qualify."
            g2_pre_advance[i].team_seasons[season_count].uni_qualifying = 1

            g1_pre_advance[i].second_pick = 3 #1 in 3 chance for second round pick
            g2_pre_advance[i].second_pick = 3


        for i in range(6):
            if g1_advance[i] in relegated[season_count]:
                relegated[season_count].remove(g1_advance[i])
            else:
                promoted[season_count].append(g1_advance[i])
            if g2_advance[i] in relegated[season_count]:
                relegated[season_count].remove(g2_advance[i])
            else:
                promoted[season_count].append(g2_advance[i])

        uni_qualif_end_time = time.time()

        with open('execution_time', 'a') as file:
            file.write(f"Season {season_count} Universal Qualifying Group Stage Execution Time: {round((uni_qualif_end_time - uni_qualif_start_time) / 60)} minutes, {round(((uni_qualif_end_time - uni_qualif_start_time) % 60), 2)} seconds.\n")

        uni_play_in_start_time = time.time()

        #UNIVERSAL PLAY IN
        advanced_play_in = []
        elim_play_in = []

        print(Fore.BLUE + "(Game 1) Group 1 no. 7 vs Group 2 no. 10, winner advances, loser eliminated")
        upi_g1W, upi_g1L = best_of(g1_advance[6], g2_advance[9],
                                   thresh=180, win_by=25,
                                   both_return=True,
                                   context=f"S{season_count} Universal Play-In Game 1",advantage=6)
        advanced_play_in.append(upi_g1W)
        elim_play_in.append(upi_g1L)
        print(Fore.GREEN + f"{upi_g1W.name} advance\n{upi_g1L.name} eliminated" + Fore.RESET)

        print(Fore.BLUE + "(Game 2) Group 2 no. 7 vs Group 1 no. 10, winner advances, loser eliminated")
        upi_g2W, upi_g2L = best_of(g2_advance[6], g1_advance[9],
                                   thresh=180, win_by=25,
                                   both_return=True,
                                   context=f"S{season_count} Universal Play-In Game 1",advantage=6)
        advanced_play_in.append(upi_g2W)
        elim_play_in.append(upi_g2L)
        print(Fore.GREEN + f"{upi_g2W.name} advance\n{upi_g2L.name} eliminated" + Fore.RESET)

        print(Fore.BLUE + "(Game 3) Group 1 no. 8 vs Group 2 no. 9, winner advances, loser eliminated")
        upi_g3W, upi_g3L = best_of(g1_advance[7], g2_advance[8],
                                   thresh=180, win_by=25,
                                   both_return=True,
                                   context=f"S{season_count} Universal Play-In Game 1",advantage=3)
        advanced_play_in.append(upi_g3W)
        elim_play_in.append(upi_g3L)
        print(Fore.GREEN + f"{upi_g3W.name} advance\n{upi_g3L.name} eliminated" + Fore.RESET)

        print(Fore.BLUE + "(Game 4) Group 2 no. 8 vs Group 1 no. 9, winner advances, loser eliminated")
        upi_g4W, upi_g4L = best_of(g2_advance[7], g1_advance[8],
                                   thresh=180, win_by=25,
                                   both_return=True,
                                   context=f"S{season_count} Universal Play-In Game 1",advantage=3)
        advanced_play_in.append(upi_g4W)
        elim_play_in.append(upi_g4L)
        print(Fore.GREEN + f"{upi_g4W.name} advance\n{upi_g4L.name} eliminated" + Fore.RESET)

        print(Fore.RED + "***END OF UNIVERSAL QUALIFYING***" + Fore.RESET)



        uni_play_in_end_time = time.time()
        with open('execution_time', 'a') as file:
            file.write(f"Season {season_count} Universal Play-In Execution Time: {round((uni_play_in_end_time - uni_play_in_start_time) / 60)} minutes, {round(((uni_play_in_end_time - uni_play_in_start_time) % 60), 2)} seconds.\n")


        advanced_play_in = list(set(advanced_play_in))

        time.sleep(3)


        for team in advanced_play_in:
            uni_teams.append(team)
            team.team_seasons[season_count].uni_qualifying = 2
            if team in relegated[season_count]:
                relegated[season_count].remove(team)
            else:
                promoted[season_count].append(team)
            team.history[season_count] += "Advanced from Universal Play-In -> UNI League. "
        for team in elim_play_in:
            team.team_seasons[season_count].uni_qualifying = 1
            team.history[season_count] += "Failed to advance from Universal Play-In. "
            void_draft.append(team)
            team.second_pick = 2 #2 in 3 chance for second round pick

        season_wipe(uni_teams)
        uni_teams = list(set(uni_teams))
        upl_standings = league_season(uni_teams, use_saved=False, season_count=season_count, region="Universal")

        for team in promoted[season_count]:
            print(f"{team.name} PROMOTED.")
            team.history[season_count] += "\n\tPROMOTED to Universal League."
        for team in relegated[season_count]:
            print(f"{team.name} RELEGATED.")
            team.history[season_count] += "\n\tRELEGATED to "

        clear_file('history')
        clear_file('my_teams')

        reverse_upl_standings = list(reversed(upl_standings))



        #to make MVP specific to regional regular season stats, DO NOT run the season_stats function on each region
        #instead, the stats will be generated in the league_season function following the regular season,
        #and the MVP can be determined by running those stats through the region_mvp function

        second_draft_g1 = []
        second_draft_g2 = []
        second_draft_g3 = []

        third_draft_full = []

        #todo currently, pre-qualifying teams are being given second round picks in a way other than being added to g3
        #every srg3 team is getting an extra second round draft pick, meaning it's being granted to them after failing
        #to qualify for the universal qualifying round
        for league in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
            for team in league:
                if team not in uni_teams:
                    all_teams.append(team)
                if team.second_pick == 1:
                    second_draft_g1.append(team)
                elif team.second_pick == 2:
                    coin = choice([True, True, False])
                    if coin:
                        team.history[season_count] += "\n\t2 in 3 chance for Second Round pick: SUCCESS"
                        second_draft_g2.append(team)
                    else:
                        team.history[season_count] += "\n\t2 in 3 chance for Second Round pick: FAILURE"
                elif team.second_pick == 3:
                    coin = choice([True, False, False])
                    if coin:
                        team.history[season_count] += "\n\t1 in 3 chance for Second Round pick: SUCCESS"
                        second_draft_g3.append(team)
                    else:
                        team.history[season_count] += "\n\t1 in 3 chance for Second Round pick: FAILURE"
                team.second_pick = 0
                if team.third_pick == 1:
                    third_draft_full.append(team)
                    team.third_pick = 0

        for team in uni_teams:
            all_teams.append(team)
            if team.second_pick == 1 and team not in second_draft_g1:
                second_draft_g1.append(team)
            elif team.second_pick == 2 and team not in second_draft_g2:
                coin = choice([True, True, False])
                if coin:
                    team.history[season_count] += "\n\t2 in 3 chance for Second Round pick: SUCCESS"
                    second_draft_g2.append(team)
                else:
                    team.history[season_count] += "\n\t2 in 3 chance for Second Round pick: FAILURE"
            elif team.second_pick == 3 and team not in second_draft_g3:
                coin = choice([True, False, False])
                if coin:
                    team.history[season_count] += "\n\t1 in 3 chance for Second Round pick: SUCCESS"
                    second_draft_g3.append(team)
                else:
                    team.history[season_count] += "\n\t1 in 3 chance for Second Round pick: FAILURE"
            if team.third_pick == 1 and team not in third_draft_full:
                third_draft_full.append(team)
                team.third_pick = 0


        shuffle(second_draft_g1)
        shuffle(second_draft_g2)
        shuffle(second_draft_g3)
        second_draft_full = second_draft_g1 + second_draft_g2 + second_draft_g3


        if True:
            all_players = []

            labyrinth_mine = 1 #I've set it equal to 1 so that I only get one Labyrinth team now
            for league in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
                league.reverse()
            for team in uni_teams:
                if team in dw_teams:
                    dw_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        dw_teams.append(add)
                        add.history[season_count] += "Darkwing Regional."
                    else:
                        add = Team('Labyrinth',season_count=season_count)
                        add.history[season_count] = "Introduced into the Darkwing Regional."
                        dw_teams.insert(0, add)
                        if labyrinth_mine <= 1:
                            add.make_mine(add.name)
                            labyrinth_mine += 1
                elif team in sc_teams:
                    sc_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        sc_teams.append(add)
                        add.history[season_count] += "Shining-Core Regional."
                    else:
                        add = Team('Labyrinth',season_count=season_count)
                        add.history[season_count] = "Introduced into the Shining-Core Regional."
                        sc_teams.insert(0, add)
                        if labyrinth_mine <= 1:
                            add.make_mine(add.name)
                            labyrinth_mine += 1
                elif team in ds_teams:
                    ds_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        ds_teams.append(add)
                        add.history[season_count] += "Diamond-Sea Regional."
                    else:
                        add = Team('Labyrinth',season_count=season_count)
                        add.history[season_count] = "Introduced into the Diamond-Sea Regional."
                        ds_teams.insert(0, add)
                        if labyrinth_mine <= 1:
                            add.make_mine(add.name)
                            labyrinth_mine += 1
                elif team in wof_teams:
                    wof_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        wof_teams.append(add)
                        add.history[season_count] += "Web-of-Nations Regional."
                    else:
                        add = Team('Labyrinth',season_count=season_count)
                        add.history[season_count] = "Introduced into the Web-of-Nations Regional."
                        wof_teams.insert(0, add)
                        if labyrinth_mine <= 1:
                            add.make_mine(add.name)
                            labyrinth_mine += 1
                elif team in iw_teams:
                    iw_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        iw_teams.append(add)
                        add.history[season_count] += "Ice-Wall Regional."
                    else:
                        add = Team('Labyrinth',season_count=season_count)
                        add.history[season_count] = "Introduced into the Ice-Wall Regional."
                        iw_teams.insert(0, add)
                        if labyrinth_mine <= 1:
                            add.make_mine(add.name)
                            labyrinth_mine += 1
                elif team in cl_teams:
                    cl_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        cl_teams.append(add)
                        add.history[season_count] += "Candyland Regional."
                    else:
                        add = Team('Labyrinth',season_count=season_count)
                        add.history[season_count] = "Introduced into the Candyland Regional."
                        cl_teams.insert(0, add)
                        if labyrinth_mine <= 1:
                            add.make_mine(add.name)
                            labyrinth_mine += 1
                elif team in hc_teams:
                    hc_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        hc_teams.append(add)
                        add.history[season_count] += "Hell's-Circle Regional."
                    else:
                        add = Team('Labyrinth',season_count=season_count)
                        add.history[season_count] = "Introduced into the Hell's-Circle Regional."
                        hc_teams.insert(0, add)
                        if labyrinth_mine <= 1:
                            add.make_mine(add.name)
                            labyrinth_mine += 1
                elif team in sh_teams:
                    sh_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        sh_teams.append(add)
                        add.history[season_count] += "Steel-Heart Regional."
                    else:
                        add = Team('Labyrinth',season_count=season_count)
                        add.history[season_count] = "Introduced into the Steel-Heart Regional."
                        sh_teams.insert(0, add)
                        if labyrinth_mine <= 1:
                            add.make_mine(add.name)
                            labyrinth_mine += 1


            with open("rosters", 'w') as l:
                l.write('')

            write_rosters = True

            total_stats = {'Power' : 0, 'Attack Damage' : 0, 'Attack Speed' : 0, 'Critical %' : 0, 'Critical X' : 0, 'Spawn Time' : 0, 'Health' : 0, 'Age' : 0}
            if write_rosters:
                with open("rosters", 'a') as w:
                    w.write(f"SEASON NO. {season_count}\n")
                    trans_region = ['Universal', 'Darkwing', 'Shining-Core', 'Diamond-Sea', 'Web-of-Nations', 'Ice-Wall', 'Candyland', "Hell's-Circle", 'Steel-Heart']
                    trans_i = -1
                    for league in [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
                        trans_i +=1
                        for team in league:
                            team.region = trans_region[trans_i]
                            for player in team.players:
                                total_stats['Power'] += player.power
                                total_stats['Attack Damage'] += player.atk_dmg
                                total_stats['Attack Speed'] += player.atk_spd
                                total_stats['Critical %'] += player.crit_pct
                                total_stats['Critical X'] += player.crit_x
                                total_stats['Spawn Time'] += player.spawn_time
                                total_stats['Health'] += player.max_health
                                total_stats['Age'] += player.age

                                all_players.append(player)
                    grade_players(all_players,is_team=True)




                    for key in total_stats.keys():
                        w.write(f"Average {key}: {total_stats[key]/len(all_players):.4f}\n")
                    w.write('\n')
                    for player in all_players:
                        w.write(f"{str(player)}\n")

            #note: reverse_upl_standings is the correct order from 1st to 36th

            other_all_teams = []
            for league in [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
                other_all_teams.extend(league)

            if season_count == 1:
                played_this_season = [
                    team for team in other_all_teams
                    if "Labyrinth" not in team.name
                ]
            else:
                played_this_season = other_all_teams

            season_wipe(played_this_season, end_of_full_season=True, season_count=season_count) #this enters player season data into SQL and sets it all to 0
            player_changes(other_all_teams, season_count=season_count)
            captain_changes(other_all_teams, season_count=season_count)
            coach_changes(other_all_teams, season_count=season_count)

            for team in reverse_upl_standings[20:]:
                void_draft.append(team)
                third_draft_full.append(team)

            void_draft.reverse()

            dw_draft = [team for team in dw_teams if team not in void_draft]
            sc_draft = [team for team in sc_teams if team not in void_draft]
            ds_draft = [team for team in ds_teams if team not in void_draft]
            wof_draft = [team for team in wof_teams if team not in void_draft]
            iw_draft = [team for team in iw_teams if team not in void_draft]
            cl_draft = [team for team in cl_teams if team not in void_draft]
            hc_draft = [team for team in hc_teams if team not in void_draft]
            sh_draft = [team for team in sh_teams if team not in void_draft]


            user_draft(dw_draft, season_count, is_regional=True, draft_name="Darkwing Regional draft")
            user_draft(sc_draft, season_count, is_regional=True, draft_name="Shining-Core Regional draft")
            user_draft(ds_draft, season_count, is_regional=True, draft_name="Diamond-Sea Regional draft")
            user_draft(wof_draft, season_count, is_regional=True, draft_name="Web-of-Nations Regional draft")
            user_draft(iw_draft, season_count,  is_regional=True, draft_name="Ice-Wall Regional draft")
            user_draft(cl_draft, season_count, is_regional=True, draft_name="Candyland Regional draft")
            user_draft(hc_draft, season_count,  is_regional=True, draft_name="Hell's-Circle Regional draft")
            user_draft(sh_draft, season_count, is_regional=True, draft_name="Steel-Heart Regional draft")

            user_draft(second_draft_full, season_count, second=True, draft_name="Secondary draft") #1 in 3 for eliminated in Universal Qualifying Group Stage
                                                                                                   #2 in 3 for eliminated Universal play-in
                                                                                                   #approx. 1 in 2 chance for regional missed playoffs
            shuffle(third_draft_full)
            user_draft(third_draft_full, season_count, third=True, draft_name="Tertiary draft") #Universal teams relegated to qualifying + 1/10 chance for regional missed playoffs

            upl_draft = reverse_upl_standings[0:20]
            upl_draft.reverse()
            #This takes the top 20 from the UPL (those NOT on the chopping block) and reverses them


            #This takes the 16 teams which were eliminated and puts the draft in order from worst to best by reversing
            #void draft includes all teams which failed to qualify from universal qualifying, as well as all teams on the universal chopping block
            upl_draft = remove_duplicates_ordered(upl_draft)


            user_draft(upl_draft, season_count, draft_name='Universal-Draft',write_draft=True)
            user_draft(void_draft, season_count, void=True, draft_name='Void-Draft')

            all_teams_lists = [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]

            def percent_within_thresholds(numbers, lim, std_dev):
                thresholds = [2, 5, 8, 10, 12, 15, std_dev, round(lim/2)]
                total = len(numbers)

                results = {
                    thresh: round((100 * sum((lim - x <= thresh and lim-x > 0) for x in list(numbers)) / total),2)
                    for thresh in thresholds
                }

                return results

            xwar_array = np.array([team.get_team_xWAR() for team_list in all_teams_lists for team in team_list])
            xwar_std = xwar_array.std()
            xwar_mean = xwar_array.mean()

            cap_xwar_array = np.array([team.captain.get_captain_xWAR() for team_list in all_teams_lists for team in team_list])
            cap_xwar_mean = cap_xwar_array.mean()

            cap = caps[season_count]
            above_cap = xwar_array[xwar_array > cap]
            below_cap = xwar_array[xwar_array <= cap]
            avg_diff_below_cap = below_cap.mean() - cap
            avg_diff_above_cap = above_cap.mean() - cap if len((list(above_cap))) > 0 else 0
            with open('history', 'w') as history:
                history.write(f"Average Team xWAR: {xwar_array.mean()}, xWAR Cap: {cap}\n")
            cap_mode = 'a' if season_count > 1 else 'w'
            with open('salary_cap_calculator', cap_mode) as cap_file:
                cap_file.write(f"S{season_count} Average Team xWAR: {xwar_mean}\nPer Captain: {cap_xwar_mean}\nPer Player: {(xwar_mean - cap_xwar_mean)/6})\nxWAR Cap: {cap}\nStd Dev: {xwar_std}\nAverage Diff Below Cap: {avg_diff_below_cap}\nAverage Diff Above Cap: {avg_diff_above_cap}\n")
                pct_thresh = percent_within_thresholds(xwar_array, cap, xwar_std)
                cap_file.write(f"% of teams OVER the cap: {round((100 * sum(x > cap for x in list(xwar_array)) / len(xwar_array)), 2)}\n")
                cap_file.write(f"% of teams UNDER the cap: {round((100 * sum(cap > x for x in list(xwar_array)) / len(xwar_array)), 2)}\n")
                for thresh in pct_thresh.keys():
                    if thresh in [2, 5, 8, 10, 12, 15]:
                        cap_file.write(f"% of teams within {thresh} xWAR of cap: {pct_thresh[thresh]}\n")
                    elif thresh == round(cap/2):
                        cap_file.write(f"% of teams over 1/2 cap value: {pct_thresh[thresh]}\n\n")
                    else:
                        cap_file.write(f"% of teams within one standard deviation of cap: {pct_thresh[thresh]}\n")

            #BIG LOOP FOR CHANGES APPLICABLE TO EVERY TEAM
            for league in [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams,
                           sh_teams]:


                for team in league:

                    #PART 1 - PRINTING HISTORY
                    team.print_team_name(season_count)
                    team.print_accolades()
                    team.print_history(season_count)
                    if team.mine:
                        print("\n" + Fore.RED + f"{team.name}\n" + Fore.BLUE + team.history[season_count] + Fore.RESET)
                    #PART 2 - SQL
                    if not ("Labyrinth" in team.name and season_count == 1):
                        team_season_params = (
                            team.team_id,
                            team.team_coach.coach_id,
                            team.name,
                            season_count,
                            team.team_seasons[season_count].region_started,
                            team.get_team_xWAR(),
                            team.team_seasons[season_count].started_universal_league,
                            team.team_seasons[season_count].regional_playoff_seed,
                            team.team_seasons[season_count].regional_final_seed,
                            team.team_seasons[season_count].last_stand,
                            team.team_seasons[season_count].pre_qualifying,
                            team.team_seasons[season_count].uni_qualifying,
                            team.team_seasons[season_count].ended_universal_league,
                            team.team_seasons[season_count].uni_playoff_seed,
                            team.team_seasons[season_count].uni_final_seed,
                            team.team_seasons[season_count].team_season_id
                        )

                        team_season_sql = """
                        UPDATE TeamSeason
                        SET
                            team_id = ?,
                            coach_id = ?,
                            team_name = ?,
                            season_count = ?,
                            region_started = ?,
                            xWAR = ?,
                            started_universal_league = ?,
                            regional_playoff_seed = ?,
                            regional_final_seed = ?,
                            last_stand = ?,
                            pre_qualifying = ?,
                            uni_qualifying = ?,
                            ended_universal_league = ?,
                            uni_playoff_seed = ?,
                            uni_final_seed = ?
                        WHERE team_season_id = ?
                        """

                        QUERY(team_season_sql, params=team_season_params, is_select=False)


                        team.team_seasons.append(TeamSeason(team))

            finalize_series_data()


            end_season_times[season_count] = time.time()
            if season_count == 1:
                print(Fore.CYAN + f"Season {season_count} Execution Time: {round((end_season_times[season_count]-start)/60)} minutes, {round(((end_season_times[season_count]-start)%60),2)} seconds." + Fore.RESET)
                with open('execution_time', 'a') as e_file:
                    e_file.write(f"Season {season_count} Total Execution Time: {round((end_season_times[season_count]-start)/60)} minutes, {round(((end_season_times[season_count]-start)%60),2)} seconds.\n\n")
            else:
                print(Fore.CYAN + f"Season {season_count} Execution Time: "
                                  f"{round((end_season_times[season_count] - end_season_times[season_count-1]) / 60)} minutes, {round(((end_season_times[season_count] - end_season_times[season_count-1]) % 60), 2)} seconds." + Fore.RESET)
                with open('execution_time', 'a') as e_file:
                    e_file.write(f"Season {season_count} Total Execution Time: "
                                  f"{round((end_season_times[season_count] - end_season_times[season_count-1]) / 60)} minutes, {round(((end_season_times[season_count] - end_season_times[season_count-1]) % 60), 2)} seconds.\n\n")

    # uni_champions.sort(key=lambda x : (x.wins / x.losses), reverse=True)
    end = time.time()
    print(f"\nTotal Execution Time: {round((end-start)/60)} minutes, {round(((end-start)%60),2)} seconds.")

def flush_database(db='C:/Users/carte/ControlDataBase.db'):
    with sqlite3.connect(db):
        pass

def test():

    teams = create_teams(26, region='Universal',season_count=1)
    for team in teams:
        for player in team.players:
            print(str(player))

def real_war_calculator():
    #the previous iteration of this was onto something, but I need to make the results independent of team systems.
    #I can accomplish this by taking a single player and calculating their rWAR in five different team situations and taking the average.
    #One iteration of this function will include a player of each tier being tested in five different team situations for a total of 20 seasons.
    replacement_player = Player(tier="R", atk_dmg=57.3, atk_spd=8, crit_pct=0.0633, crit_x=7.914, mit_pct=0.0501,
                                defense_pct=0.044, defense_abs=4, health=218.05, power=54.6, spawn_time=7,
                                crit_dmg=453.47, name="Player R")
    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    s_tier_player_x = next(p for p in real_war_team_alpha.players if p.tier == "S")

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=s_tier_player_x)

    s_tier_player_x_winrate = round(
        100 * (s_tier_player_x.game_wins / (s_tier_player_x.game_wins + s_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = s_tier_player_x.xWAR
    real_war_team_alpha.players.remove(s_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    s_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=s_tier_player_x.name)

    s_tier_rWAR_1 = s_tier_player_x_winrate - s_tier_replacement_winrate

    #S Tier, Phase 2

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    s_tier_to_remove = next(p for p in real_war_team_alpha.players if p.tier == "S") #We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(s_tier_to_remove)
    real_war_team_alpha.players.append(s_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=s_tier_player_x)

    s_tier_player_x_winrate = round(
        100 * (s_tier_player_x.game_wins / (s_tier_player_x.game_wins + s_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = s_tier_player_x.xWAR
    real_war_team_alpha.players.remove(s_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    s_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=s_tier_player_x.name)

    s_tier_rWAR_2 = s_tier_player_x_winrate - s_tier_replacement_winrate

    # S Tier, Phase 3

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    s_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "S")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(s_tier_to_remove)
    real_war_team_alpha.players.append(s_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=s_tier_player_x)

    s_tier_player_x_winrate = round(
        100 * (s_tier_player_x.game_wins / (s_tier_player_x.game_wins + s_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = s_tier_player_x.xWAR
    real_war_team_alpha.players.remove(s_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    s_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=s_tier_player_x.name)

    s_tier_rWAR_3 = s_tier_player_x_winrate - s_tier_replacement_winrate

    # S Tier, Phase 4

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    s_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "S")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(s_tier_to_remove)
    real_war_team_alpha.players.append(s_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=s_tier_player_x)

    s_tier_player_x_winrate = round(
        100 * (s_tier_player_x.game_wins / (s_tier_player_x.game_wins + s_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = s_tier_player_x.xWAR
    real_war_team_alpha.players.remove(s_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    s_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=s_tier_player_x.name)

    s_tier_rWAR_4 = s_tier_player_x_winrate - s_tier_replacement_winrate

    # S Tier, Phase 5

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    s_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "S")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(s_tier_to_remove)
    real_war_team_alpha.players.append(s_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=s_tier_player_x)

    s_tier_player_x_winrate = round(
        100 * (s_tier_player_x.game_wins / (s_tier_player_x.game_wins + s_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = s_tier_player_x.xWAR
    real_war_team_alpha.players.remove(s_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    s_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=s_tier_player_x.name)

    s_tier_rWAR_5 = s_tier_player_x_winrate - s_tier_replacement_winrate

    s_tier_rWAR_mean = mean([s_tier_rWAR_1, s_tier_rWAR_2, s_tier_rWAR_3, s_tier_rWAR_4, s_tier_rWAR_5])

    print(f"Phase 1 S Tier rWAR: {s_tier_rWAR_1:.3f}\nPhase 2 S Tier rWAR: {s_tier_rWAR_2:.3f}\nPhase 3 S Tier rWAR: {s_tier_rWAR_3:.3f}"
          f"\nPhase 4 S Tier rWAR: {s_tier_rWAR_4:.3f}\nPhase 5 S Tier rWAR: {s_tier_rWAR_5:.3f}\nMean rWAR: {s_tier_rWAR_mean:.3f}\n")


    #A Tier, Phase 1

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    a_tier_player_x = next(p for p in real_war_team_alpha.players if p.tier == "A")

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=a_tier_player_x)

    a_tier_player_x_winrate = round(
        100 * (a_tier_player_x.game_wins / (a_tier_player_x.game_wins + a_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = a_tier_player_x.xWAR
    real_war_team_alpha.players.remove(a_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    a_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=a_tier_player_x.name)

    a_tier_rWAR_1 = a_tier_player_x_winrate - a_tier_replacement_winrate

    # A Tier, Phase 2

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    a_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "A")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(a_tier_to_remove)
    real_war_team_alpha.players.append(a_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=a_tier_player_x)

    a_tier_player_x_winrate = round(
        100 * (a_tier_player_x.game_wins / (a_tier_player_x.game_wins + a_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = a_tier_player_x.xWAR
    real_war_team_alpha.players.remove(a_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    a_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=a_tier_player_x.name)

    a_tier_rWAR_2 = a_tier_player_x_winrate - a_tier_replacement_winrate

    # A Tier, Phase 3

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    a_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "A")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(a_tier_to_remove)
    real_war_team_alpha.players.append(a_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=a_tier_player_x)

    a_tier_player_x_winrate = round(
        100 * (a_tier_player_x.game_wins / (a_tier_player_x.game_wins + a_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = a_tier_player_x.xWAR
    real_war_team_alpha.players.remove(a_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    a_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=a_tier_player_x.name)

    a_tier_rWAR_3 = a_tier_player_x_winrate - a_tier_replacement_winrate

    # A Tier, Phase 4

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    a_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "A")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(a_tier_to_remove)
    real_war_team_alpha.players.append(a_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=a_tier_player_x)

    a_tier_player_x_winrate = round(
        100 * (a_tier_player_x.game_wins / (a_tier_player_x.game_wins + a_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = a_tier_player_x.xWAR
    real_war_team_alpha.players.remove(a_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    a_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=a_tier_player_x.name)

    a_tier_rWAR_4 = a_tier_player_x_winrate - a_tier_replacement_winrate

    # A Tier, Phase 5

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    a_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "A")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(a_tier_to_remove)
    real_war_team_alpha.players.append(a_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=a_tier_player_x)

    a_tier_player_x_winrate = round(
        100 * (a_tier_player_x.game_wins / (a_tier_player_x.game_wins + a_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = a_tier_player_x.xWAR
    real_war_team_alpha.players.remove(a_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    a_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=a_tier_player_x.name)

    a_tier_rWAR_5 = a_tier_player_x_winrate - a_tier_replacement_winrate

    a_tier_rWAR_mean = mean([a_tier_rWAR_1, a_tier_rWAR_2, a_tier_rWAR_3, a_tier_rWAR_4, a_tier_rWAR_5])

    print(
        f"Phase 1 A Tier rWAR: {a_tier_rWAR_1:.3f}\nPhase 2 A Tier rWAR: {a_tier_rWAR_2:.3f}\nPhase 3 A Tier rWAR: {a_tier_rWAR_3:.3f}"
        f"\nPhase 4 A Tier rWAR: {a_tier_rWAR_4:.3f}\nPhase 5 A Tier rWAR: {a_tier_rWAR_5:.3f}\nMean rWAR: {a_tier_rWAR_mean:.3f}\n")


    #B Tier, Phase 1

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    b_tier_player_x = next(p for p in real_war_team_alpha.players if p.tier == "B")

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=b_tier_player_x)

    b_tier_player_x_winrate = round(
        100 * (b_tier_player_x.game_wins / (b_tier_player_x.game_wins + b_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = b_tier_player_x.xWAR
    real_war_team_alpha.players.remove(b_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    b_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=b_tier_player_x.name)

    b_tier_rWAR_1 = b_tier_player_x_winrate - b_tier_replacement_winrate

    #B Tier, Phase 2

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    b_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "B")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(b_tier_to_remove)
    real_war_team_alpha.players.append(b_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=b_tier_player_x)

    b_tier_player_x_winrate = round(
        100 * (b_tier_player_x.game_wins / (b_tier_player_x.game_wins + b_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = b_tier_player_x.xWAR
    real_war_team_alpha.players.remove(b_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    b_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=b_tier_player_x.name)

    b_tier_rWAR_2 = b_tier_player_x_winrate - b_tier_replacement_winrate

    #B Tier, Phase 3

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    b_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "B")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(b_tier_to_remove)
    real_war_team_alpha.players.append(b_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=b_tier_player_x)

    b_tier_player_x_winrate = round(
        100 * (b_tier_player_x.game_wins / (b_tier_player_x.game_wins + b_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = b_tier_player_x.xWAR
    real_war_team_alpha.players.remove(b_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    b_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=b_tier_player_x.name)

    b_tier_rWAR_3 = b_tier_player_x_winrate - b_tier_replacement_winrate

    #B Tier, Phase 4

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    b_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "B")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(b_tier_to_remove)
    real_war_team_alpha.players.append(b_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=b_tier_player_x)

    b_tier_player_x_winrate = round(
        100 * (b_tier_player_x.game_wins / (b_tier_player_x.game_wins + b_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = b_tier_player_x.xWAR
    real_war_team_alpha.players.remove(b_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    b_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=b_tier_player_x.name)

    b_tier_rWAR_4 = b_tier_player_x_winrate - b_tier_replacement_winrate

    #B Tier, Phase 5

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    b_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "B")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(b_tier_to_remove)
    real_war_team_alpha.players.append(b_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=b_tier_player_x)

    b_tier_player_x_winrate = round(
        100 * (b_tier_player_x.game_wins / (b_tier_player_x.game_wins + b_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = b_tier_player_x.xWAR
    real_war_team_alpha.players.remove(b_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    b_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=b_tier_player_x.name)

    b_tier_rWAR_5 = b_tier_player_x_winrate - b_tier_replacement_winrate

    b_tier_rWAR_mean = mean([b_tier_rWAR_1, b_tier_rWAR_2, b_tier_rWAR_3, b_tier_rWAR_4, b_tier_rWAR_5])

    print(
        f"Phase 1 B Tier rWAR: {b_tier_rWAR_1:.3f}\nPhase 2 B Tier rWAR: {b_tier_rWAR_2:.3f}\nPhase 3 B Tier rWAR: {b_tier_rWAR_3:.3f}"
        f"\nPhase 4 B Tier rWAR: {b_tier_rWAR_4:.3f}\nPhase 5 B Tier rWAR: {b_tier_rWAR_5:.3f}\nMean rWAR: {b_tier_rWAR_mean:.3f}\n")

    # C Tier, Phase 1

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_player_x = next(p for p in real_war_team_alpha.players if p.tier == "C")

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=c_tier_player_x)

    c_tier_player_x_winrate = round(
        100 * (c_tier_player_x.game_wins / (c_tier_player_x.game_wins + c_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = c_tier_player_x.xWAR
    real_war_team_alpha.players.remove(c_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    c_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=c_tier_player_x.name)

    c_tier_rWAR_1 = c_tier_player_x_winrate - c_tier_replacement_winrate

    # C Tier, Phase 2

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(c_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=c_tier_player_x)

    c_tier_player_x_winrate = round(
        100 * (c_tier_player_x.game_wins / (c_tier_player_x.game_wins + c_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = c_tier_player_x.xWAR
    real_war_team_alpha.players.remove(c_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    c_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=c_tier_player_x.name)

    c_tier_rWAR_2 = c_tier_player_x_winrate - c_tier_replacement_winrate

    # C Tier, Phase 3

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(c_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=c_tier_player_x)

    c_tier_player_x_winrate = round(
        100 * (c_tier_player_x.game_wins / (c_tier_player_x.game_wins + c_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = c_tier_player_x.xWAR
    real_war_team_alpha.players.remove(c_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    c_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=c_tier_player_x.name)

    c_tier_rWAR_3 = c_tier_player_x_winrate - c_tier_replacement_winrate

    # C Tier, Phase 4

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(c_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=c_tier_player_x)

    c_tier_player_x_winrate = round(
        100 * (c_tier_player_x.game_wins / (c_tier_player_x.game_wins + c_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = c_tier_player_x.xWAR
    real_war_team_alpha.players.remove(c_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    c_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=c_tier_player_x.name)

    c_tier_rWAR_4 = c_tier_player_x_winrate - c_tier_replacement_winrate

    # C Tier, Phase 5

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the S Tier Player X
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(c_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=c_tier_player_x)

    c_tier_player_x_winrate = round(
        100 * (c_tier_player_x.game_wins / (c_tier_player_x.game_wins + c_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = c_tier_player_x.xWAR
    real_war_team_alpha.players.remove(c_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    c_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=c_tier_player_x.name)

    c_tier_rWAR_5 = c_tier_player_x_winrate - c_tier_replacement_winrate

    c_tier_rWAR_mean = mean([c_tier_rWAR_1, c_tier_rWAR_2, c_tier_rWAR_3, c_tier_rWAR_4, c_tier_rWAR_5])

    print(
        f"Phase 1 C Tier rWAR: {c_tier_rWAR_1:.3f}\nPhase 2 C Tier rWAR: {c_tier_rWAR_2:.3f}\nPhase 3 C Tier rWAR: {c_tier_rWAR_3:.3f}"
        f"\nPhase 4 C Tier rWAR: {c_tier_rWAR_4:.3f}\nPhase 5 C Tier rWAR: {c_tier_rWAR_5:.3f}\nMean rWAR: {c_tier_rWAR_mean:.3f}\n")


    #D Tier, Phase 1

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)


    d_tier_crit_x=round(uniform(6.3,7.75), 3)
    d_tier_atk_dmg = round(uniform(48,57), 1)
    d_tier_crit_dmg = d_tier_crit_x*d_tier_atk_dmg
    d_tier_player_x = Player(tier="D", atk_dmg=d_tier_atk_dmg, atk_spd=choice([8,9]), crit_pct=round(uniform(0.05,0.06), 4),
                            crit_x=d_tier_crit_x, mit_pct=round(uniform(0.015,0.049), 4),
                            defense_pct=round(uniform(0.029,0.04), 4), defense_abs=choice([2,3,4]), health=round(uniform(209,217), 2), power=choice([51,52,53,54]), spawn_time=choice([7,8]),
                            crit_dmg=d_tier_crit_dmg, name="D Tier Test")

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the "D" Tier Player
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(d_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=d_tier_player_x)

    d_tier_player_x_winrate = round(
        100 * (d_tier_player_x.game_wins / (d_tier_player_x.game_wins + d_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = d_tier_player_x.xWAR
    real_war_team_alpha.players.remove(d_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    d_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=d_tier_player_x.name)

    d_tier_rWAR_1 = d_tier_player_x_winrate - d_tier_replacement_winrate

    # D Tier, Phase 2

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the "D" Tier Player
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(d_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=d_tier_player_x)

    d_tier_player_x_winrate = round(
        100 * (d_tier_player_x.game_wins / (d_tier_player_x.game_wins + d_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = d_tier_player_x.xWAR
    real_war_team_alpha.players.remove(d_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    d_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=d_tier_player_x.name)

    d_tier_rWAR_2 = d_tier_player_x_winrate - d_tier_replacement_winrate

    # D Tier, Phase 3

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the "D" Tier Player
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(d_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=d_tier_player_x)

    d_tier_player_x_winrate = round(
        100 * (d_tier_player_x.game_wins / (d_tier_player_x.game_wins + d_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = d_tier_player_x.xWAR
    real_war_team_alpha.players.remove(d_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    d_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=d_tier_player_x.name)

    d_tier_rWAR_3 = d_tier_player_x_winrate - d_tier_replacement_winrate

    # D Tier, Phase 4

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the "D" Tier Player
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(d_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=d_tier_player_x)

    d_tier_player_x_winrate = round(
        100 * (d_tier_player_x.game_wins / (d_tier_player_x.game_wins + d_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = d_tier_player_x.xWAR
    real_war_team_alpha.players.remove(d_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    d_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=d_tier_player_x.name)

    d_tier_rWAR_4 = d_tier_player_x_winrate - d_tier_replacement_winrate

    # D Tier, Phase 5

    TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
    real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
    TEAMS.insert(0, real_war_team_alpha)

    c_tier_to_remove = next(
        p for p in real_war_team_alpha.players if p.tier == "C")  # We will replace this player with the "D" Tier Player
    real_war_team_alpha.players.remove(c_tier_to_remove)
    real_war_team_alpha.players.append(d_tier_player_x)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=d_tier_player_x)

    d_tier_player_x_winrate = round(
        100 * (d_tier_player_x.game_wins / (d_tier_player_x.game_wins + d_tier_player_x.game_losses)), 2)

    replacement_player.xWAR = d_tier_player_x.xWAR
    real_war_team_alpha.players.remove(d_tier_player_x)
    real_war_team_alpha.players.append(replacement_player)
    real_war_team_alpha.lineups = generate_lineups_six_to_four(real_war_team_alpha.players,
                                                               real_war_team_alpha.team_coach,
                                                               real_war_team_alpha.team_id)

    d_tier_replacement_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                             replacement_player=replacement_player, replacing=d_tier_player_x.name)

    d_tier_rWAR_5 = d_tier_player_x_winrate - d_tier_replacement_winrate

    d_tier_rWAR_mean = mean([d_tier_rWAR_1, d_tier_rWAR_2, d_tier_rWAR_3, d_tier_rWAR_4, d_tier_rWAR_5])

    print(
        f"Phase 1 D Tier rWAR: {d_tier_rWAR_1:.3f}\nPhase 2 D Tier rWAR: {d_tier_rWAR_2:.3f}\nPhase 3 D Tier rWAR: {d_tier_rWAR_3:.3f}"
        f"\nPhase 4 D Tier rWAR: {d_tier_rWAR_4:.3f}\nPhase 5 D Tier rWAR: {d_tier_rWAR_5:.3f}\nMean rWAR: {d_tier_rWAR_mean:.3f}\n")




    s_tier_rWAR_mean = round(s_tier_rWAR_mean, 3)
    a_tier_rWAR_mean = round(a_tier_rWAR_mean, 3)
    b_tier_rWAR_mean = round(b_tier_rWAR_mean, 3)
    c_tier_rWAR_mean = round(c_tier_rWAR_mean, 3)
    d_tier_rWAR_mean = round(d_tier_rWAR_mean, 3)




    rows = [

        {"Name": s_tier_player_x.name, "xWAR": s_tier_player_x.xWAR, "rWAR": s_tier_rWAR_mean,
         "Attack Damage" : s_tier_player_x.atk_dmg, "Attack Speed" : s_tier_player_x.atk_spd,
         "Critical Chance" : s_tier_player_x.crit_pct, "Critical Multiplier" : s_tier_player_x.crit_x,
         "Defense %" : s_tier_player_x.defense_pct, "Defense Absolute" : s_tier_player_x.defense_abs,
         "Mitigation Chance" : s_tier_player_x.mit_pct, "Max Health" : s_tier_player_x.max_health,
         "Power" : s_tier_player_x.power, "Spawn Time" : s_tier_player_x.spawn_time,
         "Primary Trait" : s_tier_player_x.trait_tag[0], "Secondary Trait" : s_tier_player_x.trait_tag[1]},

        {"Name": a_tier_player_x.name, "xWAR": a_tier_player_x.xWAR, "rWAR": a_tier_rWAR_mean,
         "Attack Damage": a_tier_player_x.atk_dmg, "Attack Speed": a_tier_player_x.atk_spd,
         "Critical Chance": a_tier_player_x.crit_pct, "Critical Multiplier": a_tier_player_x.crit_x,
         "Defense %": a_tier_player_x.defense_pct, "Defense Absolute": a_tier_player_x.defense_abs,
         "Mitigation Chance": a_tier_player_x.mit_pct, "Max Health": a_tier_player_x.max_health,
         "Power": a_tier_player_x.power, "Spawn Time": a_tier_player_x.spawn_time,
         "Primary Trait": a_tier_player_x.trait_tag[0], "Secondary Trait": a_tier_player_x.trait_tag[1]},

        {"Name": b_tier_player_x.name, "xWAR": b_tier_player_x.xWAR, "rWAR": b_tier_rWAR_mean,
         "Attack Damage": b_tier_player_x.atk_dmg, "Attack Speed": b_tier_player_x.atk_spd,
         "Critical Chance": b_tier_player_x.crit_pct, "Critical Multiplier": b_tier_player_x.crit_x,
         "Defense %": b_tier_player_x.defense_pct, "Defense Absolute": b_tier_player_x.defense_abs,
         "Mitigation Chance": b_tier_player_x.mit_pct, "Max Health": b_tier_player_x.max_health,
         "Power": b_tier_player_x.power, "Spawn Time": b_tier_player_x.spawn_time,
         "Primary Trait": b_tier_player_x.trait_tag[0], "Secondary Trait": b_tier_player_x.trait_tag[1]},

        {"Name": c_tier_player_x.name, "xWAR": c_tier_player_x.xWAR, "rWAR": c_tier_rWAR_mean,
         "Attack Damage": c_tier_player_x.atk_dmg, "Attack Speed": c_tier_player_x.atk_spd,
         "Critical Chance": c_tier_player_x.crit_pct, "Critical Multiplier": c_tier_player_x.crit_x,
         "Defense %": c_tier_player_x.defense_pct, "Defense Absolute": c_tier_player_x.defense_abs,
         "Mitigation Chance": c_tier_player_x.mit_pct, "Max Health": c_tier_player_x.max_health,
         "Power": c_tier_player_x.power, "Spawn Time": c_tier_player_x.spawn_time,
         "Primary Trait": c_tier_player_x.trait_tag[0], "Secondary Trait": c_tier_player_x.trait_tag[1]},

        {"Name": d_tier_player_x.name, "xWAR": d_tier_player_x.xWAR, "rWAR": d_tier_rWAR_mean,
         "Attack Damage": d_tier_player_x.atk_dmg, "Attack Speed": d_tier_player_x.atk_spd,
         "Critical Chance": d_tier_player_x.crit_pct, "Critical Multiplier": d_tier_player_x.crit_x,
         "Defense %": d_tier_player_x.defense_pct, "Defense Absolute": d_tier_player_x.defense_abs,
         "Mitigation Chance": d_tier_player_x.mit_pct, "Max Health": d_tier_player_x.max_health,
         "Power": d_tier_player_x.power, "Spawn Time": d_tier_player_x.spawn_time,
         "Primary Trait": d_tier_player_x.trait_tag[0], "Secondary Trait": d_tier_player_x.trait_tag[1]}

    ]

    file_path = "rWAR Data.xlsx"

    # try to load existing data
    try:
        existing_df = pd.read_excel(file_path)
    except FileNotFoundError:
        existing_df = pd.DataFrame()

    # new data
    new_df = pd.DataFrame(rows)

    # append
    combined_df = pd.concat([existing_df, new_df], ignore_index=True)

    # write back
    combined_df.to_excel(file_path, index=False)


#Average Captain:
#damage_taken = 0.102167
#max_health = 95.79
#atk_dmg_bonus = 1.17618
#crit_pct_bonus = 0
#crit_x_bonus = 1.37442
#power_bonus = 0.46402

def captain_rWAR_calculator(rounds):
    replacement_captain = Captain()
    replacement_captain.name = "Captain R"
    replacement_captain.damage_taken = 0.102167
    replacement_captain.max_health = 95.79
    replacement_captain.atk_dmg_bonus = 1.17618
    replacement_captain.crit_x_bonus = 1.37442
    replacement_captain.power_bonus = 0.46402

    non_captain = Captain()
    non_captain.name = "Non-Captain"
    non_captain.damage_taken = 0
    non_captain.max_health = 0
    non_captain.atk_dmg_bonus = 1
    non_captain.crit_x_bonus = 1
    non_captain.power_bonus = 0

    crazy_captain = Captain()
    crazy_captain.name = "Immortal Captain w/ normal bonuses"
    crazy_captain.damage_taken = 0
    crazy_captain.max_health = 1
    crazy_captain.atk_dmg_bonus = 1.25
    crazy_captain.crit_x_bonus = 1.5
    crazy_captain.power_bonus = 0.5



    for _ in range(rounds):

        #PHASE 1

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        baseline_captain = real_war_team_alpha.captain #This captain will be used for all five phases

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=baseline_captain)

        #real_war_team_alpha.captain = crazy_captain

        #crazy_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=crazy_captain, replacing=baseline_captain.name)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=replacement_captain, replacing=baseline_captain.name)

        #real_war_team_alpha.captain = non_captain

        #non_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=non_captain, replacing=baseline_captain.name)

        captain_rWAR1 = round((baseline_captain_winrate-replacement_captain_winrate), 2)

        #PHASE 2

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha, replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR2 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        # PHASE 3

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                               replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR3 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        # PHASE 4

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                               replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR4 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        # PHASE 5

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                               replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR5 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        # PHASE 6

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                               replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR6 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        # PHASE 7

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                               replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR7 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        # PHASE 8

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                               replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR8 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        # PHASE 9

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                               replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR9 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        # PHASE 10

        TEAMS = create_teams(count=50, region="Test", draft=False, allow_duplicate_names=True)
        real_war_team_alpha = Team(region="Test", mine=True, pre_name="Team Alpha", season_count=0)
        TEAMS.insert(0, real_war_team_alpha)

        real_war_team_alpha.captain = baseline_captain

        baseline_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                               replacement_player=baseline_captain)

        real_war_team_alpha.captain = replacement_captain

        replacement_captain_winrate = round_robin(TEAMS, 10, 0, 4, real_war_team_alpha=real_war_team_alpha,
                                                  replacement_player=replacement_captain,
                                                  replacing=baseline_captain.name)

        captain_rWAR10 = round((baseline_captain_winrate - replacement_captain_winrate), 2)

        captain_rWAR_mean = round(mean([captain_rWAR1, captain_rWAR2, captain_rWAR3, captain_rWAR4, captain_rWAR5, captain_rWAR6, captain_rWAR7, captain_rWAR8, captain_rWAR9, captain_rWAR10]), 3)

        row = [

            {"Name" : baseline_captain.name, "rWAR" : captain_rWAR_mean,
             "Damage Taken": baseline_captain.damage_taken,
             "Max Health" : baseline_captain.max_health, "Base Damage Bonus" : baseline_captain.atk_dmg_bonus,
             "Critical Damage Bonus" : baseline_captain.crit_x_bonus, "Power Bonus" : baseline_captain.power_bonus}

        ]

        file_path = "Captain rWAR Data.xlsx"

        # try to load existing data
        try:
            existing_df = pd.read_excel(file_path)
        except FileNotFoundError:
            existing_df = pd.DataFrame()

        # new data
        new_df = pd.DataFrame(row)

        # append
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)

        # write back
        combined_df.to_excel(file_path, index=False)


def collect_data():
    for i in range(500):
        real_war_calculator()
        if i % 2 == 0:
            captain_rWAR_calculator(2)
        else:
            captain_rWAR_calculator(3)



    #I need to upgrade this to give every single player 6 different rWAR values calculated in 6 different teams and against
#6 different sets of opponents. These can be rWAR1 through rWAR6 and their final rWAR
#It is also a problem that I am entering data for players on the same team
#Instead of a phase system that goes through a whole team, I need to take one player from a team, find their winrate and subtract it by the winrate of a replacement player in the same slot.
#Store this as rWAR1. Run it again with the same player but a totally new team and set of 50 opponents. This difference will be rWAR2. Do the same thing four more times and store each value as a new rWAR.
#The average value of the 6 rWAR values is essentially their overall value and the two extremes show their potential. High variance players are more dependent on their systems to succeed, while low variance players
#are more consistent. Perhaps I can call the final metric uWAR which takes the average rWAR and penalizes players with high variance.

#clear_all_databases()
#initiate_databases()
#collect_data()
main()

#for _ in range(1000):
#    real_war_calculator()

